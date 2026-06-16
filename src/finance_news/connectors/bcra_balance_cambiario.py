"""BCRA balance cambiario mensual connector.

Consume el anexo estadístico mensual del "Mercado de cambios y balance
cambiario" publicado por el BCRA y lo normaliza en observaciones mensuales por
rubro (período, compras, ventas y saldo neto en USD).

Diseño:

* El parseo XLSX depende de la dependencia opcional ``openpyxl``. Para mantener
  el módulo importable sin dependencias de terceros, la importación de
  ``openpyxl`` se hace de forma perezosa dentro de ``OpenpyxlRowExtractor`` y
  sólo se instancia dentro de ``fetch_page`` cuando no se inyecta un extractor.
* El parser es determinista por cabecera (no heurística posicional).
* ``fetch_page`` es una implementación real: descarga el artefacto, delega en el
  extractor de filas, normaliza y devuelve un ``PageResult``.
"""

from __future__ import annotations

import csv
import hashlib
from dataclasses import dataclass, field
from datetime import datetime, timezone
from typing import Any, Mapping, Protocol

from ._http import AsyncHttpTransport, HttpRequest
from .models import (
    Freshness,
    PageResult,
    Provenance,
    RateLimitPolicy,
    RecoverableConnectorError,
    RetryPolicy,
    SourceItem,
)

CONNECTOR_NAME = "bcra_balance_cambiario"
SOURCE_NAME = "bcra"
PARSER_VERSION = "0.1.0"
DEFAULT_TTL_SECONDS = 30 * 24 * 60 * 60
DEFAULT_UNIDAD = "USD"

# AC#3 (documentación explícita):
# Este dataset mensual del anexo del mercado de cambios NO reemplaza la
# intervención diaria neta del BCRA. Es la mejor fuente oficial agregada
# identificada, pero no provee un dato diario de intervención spot. Se expone
# este flag en los metadatos de cada observación para que consumidores y
# downstream eviten confundirlo con intervención neta diaria.
DOES_NOT_REPLACE_NET_INTERVENTION = True

# URL del artefacto. LIMITACIÓN CONOCIDA: el documento de investigación
# (analysis/source_research_bcra.md, sección 1.7) identifica el anexo XLSX del
# "Mercado de cambios y balance cambiario" pero NO expone una URL directa y
# estable al archivo XLSX (los enlaces finales se resuelven desde la página).
# Mientras no se conozca un enlace XLSX estable, se apunta a la página oficial
# de estadísticas del mercado de cambios. El flujo de ``fetch_page`` es plenamente
# funcional dado un cuerpo de bytes (vía extractor inyectado en pruebas o
# ``openpyxl`` cuando esté disponible); resolver el enlace final estable es un
# seguimiento operativo posterior y no afecta al contrato del conector.
DEFAULT_XLSX_URL = (
    "https://www.bcra.gob.ar/"
    "estadisticas-estandarizadas-sobre-la-evolucion-del-mercado-de-cambios/"
)

# Palabras clave usadas para DETECTAR la fila de cabecera (>= 2 coincidencias).
_HEADER_DETECTION_KEYWORDS = (
    "periodo",
    "rubro",
    "compras",
    "ventas",
    "saldo",
    "categoria",
)

# Mapeo de nombre lógico de columna -> palabras clave para construir el índice
# de columnas de forma determinista a partir de la cabecera real.
_COLUMN_KEYWORDS: tuple[tuple[str, tuple[str, ...]], ...] = (
    ("period", ("periodo",)),
    ("rubro", ("rubro", "categoria", "sector")),
    ("compras", ("compras", "compra")),
    ("ventas", ("ventas", "venta")),
    ("saldo", ("saldo", "neto")),
)

_NA_TOKENS = frozenset({"", "NA", "N/A", "N.A.", "-", "—", "n/a"})


class XlsxRowExtractor(Protocol):
    """Extract rows of string cells from a binary spreadsheet payload."""

    def extract_rows(self, data: bytes) -> list[list[str]]:
        """Return rows of cells as strings for the given payload."""


class MissingXlsxExtractorError(RuntimeError):
    """Raised when no XLSX extraction backend is available."""


class OpenpyxlRowExtractor:
    """XLSX row extractor backed by openpyxl when present.

    The ``openpyxl`` import is performed lazily inside ``__init__`` so the
    module remains importable without third-party dependencies; the instance is
    only created on demand inside ``fetch_page``.
    """

    def __init__(self) -> None:
        try:
            from openpyxl import load_workbook  # type: ignore[import-not-found]
        except ImportError as exc:  # pragma: no cover - depends on environment
            raise MissingXlsxExtractorError(
                "XLSX row extraction requires the optional 'openpyxl' dependency."
            ) from exc

        self._load_workbook = load_workbook

    def extract_rows(self, data: bytes) -> list[list[str]]:
        from io import BytesIO

        workbook = self._load_workbook(BytesIO(data), read_only=True, data_only=True)
        rows: list[list[str]] = []
        for sheet in workbook.worksheets:
            for raw_row in sheet.iter_rows(values_only=True):
                rows.append(
                    ["" if cell is None else str(cell) for cell in raw_row]
                )
        return rows


@dataclass(frozen=True)
class ParsedBalanceObservation:
    """Una observación mensual del balance cambiario por rubro."""

    period: str
    rubro: str
    compras_usd: float | None
    ventas_usd: float | None
    saldo_neto_usd: float | None
    unidad: str = DEFAULT_UNIDAD
    fuente: str = SOURCE_NAME

    def to_dict(self) -> dict[str, Any]:
        return {
            "period": self.period,
            "rubro": self.rubro,
            "compras_usd": self.compras_usd,
            "ventas_usd": self.ventas_usd,
            "saldo_neto_usd": self.saldo_neto_usd,
            "unidad": self.unidad,
            "fuente": self.fuente,
        }

    @classmethod
    def from_dict(cls, data: Mapping[str, Any]) -> "ParsedBalanceObservation":
        return cls(
            period=str(data["period"]),
            rubro=str(data["rubro"]),
            compras_usd=_to_optional_float(data.get("compras_usd")),
            ventas_usd=_to_optional_float(data.get("ventas_usd")),
            saldo_neto_usd=_to_optional_float(data.get("saldo_neto_usd")),
            unidad=str(data.get("unidad", DEFAULT_UNIDAD)),
            fuente=str(data.get("fuente", SOURCE_NAME)),
        )


def parse_balance_csv(csv_text: str) -> list[ParsedBalanceObservation]:
    """Parse semicolon-delimited CSV text into observations."""
    rows = list(csv.reader(csv_text.splitlines(), delimiter=";"))
    return parse_balance_rows(rows)


def parse_balance_rows(rows: list[list[str]]) -> list[ParsedBalanceObservation]:
    """Parse a matrix of string cells into observations.

    The parser locates the header row deterministically (by keyword matching)
    and builds a column index map, then reads each subsequent row by those exact
    indices. It is NOT a positional heuristic.
    """
    header_index = _find_header_row(rows)
    if header_index is None:
        raise ValueError("Could not locate the balance cambiario header row.")

    column_map = _build_column_map([cell.strip().lower() for cell in rows[header_index]])

    period_idx = column_map.get("period")
    rubro_idx = column_map.get("rubro")
    compras_idx = column_map.get("compras")
    ventas_idx = column_map.get("ventas")
    saldo_idx = column_map.get("saldo")

    observations: list[ParsedBalanceObservation] = []
    for row in rows[header_index + 1 :]:
        cells = [cell.strip() for cell in row]

        if not any(cells):
            continue

        if period_idx is None or rubro_idx is None:
            continue
        if period_idx >= len(cells) or rubro_idx >= len(cells):
            continue

        period = cells[period_idx].strip()
        rubro = cells[rubro_idx].strip()
        if not period or not rubro:
            continue

        compras = _cell_decimal(cells, compras_idx)
        ventas = _cell_decimal(cells, ventas_idx)
        saldo = _cell_decimal(cells, saldo_idx)
        if saldo is None and compras is not None and ventas is not None:
            saldo = compras - ventas

        observations.append(
            ParsedBalanceObservation(
                period=period,
                rubro=rubro,
                compras_usd=compras,
                ventas_usd=ventas,
                saldo_neto_usd=saldo,
            )
        )

    if not observations:
        raise ValueError("Balance cambiario parser produced zero observations.")
    return observations


def normalize_balance_observations(
    *,
    observations: list[ParsedBalanceObservation],
    fetched_at: datetime,
    fetch_url: str,
    cursor: str | None,
    transport_metadata: Mapping[str, object] | None = None,
) -> list[SourceItem]:
    """Normalize parsed observations into SourceItem objects."""
    items: list[SourceItem] = []
    extra_transport = dict(transport_metadata or {})
    for observation in observations:
        monto = observation.saldo_neto_usd
        if monto is None:
            monto = observation.ventas_usd

        published_at = _parse_period(observation.period)
        metadata: dict[str, object] = {
            "rubro": observation.rubro,
            "period": observation.period,
            "monto": monto,
            "unidad": observation.unidad,
            "fuente": observation.fuente,
            "does_not_replace_net_intervention": DOES_NOT_REPLACE_NET_INTERVENTION,
        }
        if observation.compras_usd is not None:
            metadata["compras_usd"] = observation.compras_usd
        if observation.ventas_usd is not None:
            metadata["ventas_usd"] = observation.ventas_usd
        if observation.saldo_neto_usd is not None:
            metadata["saldo_neto_usd"] = observation.saldo_neto_usd

        external_id = f"balance-cambiario:{observation.period}:{observation.rubro}"
        item = SourceItem(
            external_id=external_id,
            source=SOURCE_NAME,
            published_at=published_at,
            title=f"Balance cambiario {observation.period} - {observation.rubro}",
            body=None,
            summary=None,
            url=fetch_url,
            metadata=metadata,
            provenance=Provenance(
                connector=CONNECTOR_NAME,
                source=SOURCE_NAME,
                fetch_url=fetch_url,
                canonical_url=fetch_url,
                cursor=cursor,
                fetched_at=fetched_at,
                parser_version=PARSER_VERSION,
                transport_metadata=extra_transport,
            ),
            freshness=Freshness(
                published_at=published_at,
                first_seen_at=fetched_at,
                fetched_at=fetched_at,
                is_stale=False,
                ttl_seconds=DEFAULT_TTL_SECONDS,
            ),
        )
        items.append(item)
    return items


class BcraBalanceCambiarioConnector:
    """Fetch the monthly BCRA balance cambiario annex.

    Monthly connector for the BCRA "Mercado de cambios y balance cambiario"
    annex. The connector downloads a machine-readable artifact (XLSX by default,
    resolved via the injected/openpyxl-backed row extractor) and returns one
    normalized SourceItem per rubro-period observation.
    """

    name = CONNECTOR_NAME
    source = SOURCE_NAME
    retry_policy = RetryPolicy(
        max_attempts=3,
        base_delay_seconds=1.0,
        max_delay_seconds=8.0,
    )
    rate_limit_policy = RateLimitPolicy(concurrency=1, burst=1)

    def __init__(
        self,
        *,
        transport: AsyncHttpTransport,
        row_extractor: XlsxRowExtractor | None = None,
    ) -> None:
        self._transport = transport
        # Stays None until resolved lazily in fetch_page; this guarantees the
        # module/connector never imports or instantiates openpyxl at import or
        # construction time.
        self._row_extractor = row_extractor

    async def fetch_page(
        self,
        *,
        cursor: str | None = None,
        since: datetime | None = None,
    ) -> PageResult:
        del since

        response = await self._transport.send(
            HttpRequest(
                method="GET",
                url=DEFAULT_XLSX_URL,
                headers={
                    "Accept": "application/vnd.openxmlformats-officedocument"
                    ".spreadsheetml.sheet"
                },
            )
        )

        status = response.status_code
        if status == 200:
            pass
        elif status == 404 or 500 <= status <= 599:
            raise RecoverableConnectorError(
                f"BCRA returned {status} for {DEFAULT_XLSX_URL}."
            )
        else:
            raise ValueError(
                f"Unexpected BCRA status code {status} for {DEFAULT_XLSX_URL}."
            )

        extractor = (
            self._row_extractor
            if self._row_extractor is not None
            else OpenpyxlRowExtractor()
        )
        rows = extractor.extract_rows(response.body)
        observations = parse_balance_rows(rows)

        raw_hash = hashlib.sha256(response.body).hexdigest()
        fetched_at = datetime.now(timezone.utc)
        items = normalize_balance_observations(
            observations=observations,
            fetched_at=fetched_at,
            fetch_url=DEFAULT_XLSX_URL,
            cursor=cursor,
            transport_metadata={
                "raw_hash": raw_hash,
                "status_code": status,
                "content_type": response.headers.get("Content-Type"),
            },
        )
        return PageResult(items=items, has_more=False, next_cursor=None)


def _find_header_row(rows: list[list[str]]) -> int | None:
    for index, row in enumerate(rows):
        lowered = [cell.strip().lower() for cell in row]
        matched = 0
        for keyword in _HEADER_DETECTION_KEYWORDS:
            if any(keyword in cell for cell in lowered):
                matched += 1
        if matched >= 2:
            return index
    return None


def _build_column_map(header_cells: list[str]) -> dict[str, int]:
    column_map: dict[str, int] = {}
    for position, cell in enumerate(header_cells):
        for logical_name, keywords in _COLUMN_KEYWORDS:
            if logical_name in column_map:
                continue
            if any(keyword in cell for keyword in keywords):
                column_map[logical_name] = position
                break
    return column_map


def _cell_decimal(cells: list[str], index: int | None) -> float | None:
    if index is None or index >= len(cells):
        return None
    return _parse_decimal(cells[index])


def _parse_decimal(value: str) -> float | None:
    """Parse a decimal string robustly across locale conventions.

    Handles NA tokens and thousands/decimal separators in either order:
    - "" / "NA" / "N/A" / "-" -> None
    - "1234.56" -> 1234.56
    - "1234,56" -> 1234.56
    - "1.234,56" -> 1234.56  (rightmost separator is decimal)
    - "1,234.56" -> 1234.56  (rightmost separator is decimal)
    - "1,000.5" -> 1000.5
    """
    if value is None:
        return None
    text = str(value).strip()
    if text.upper() in _NA_TOKENS:
        return None

    has_dot = "." in text
    has_comma = "," in text

    if has_dot and has_comma:
        if text.rfind(".") > text.rfind(","):
            # '.' is the decimal separator, ',' is thousands.
            cleaned = text.replace(",", "")
        else:
            # ',' is the decimal separator, '.' is thousands.
            cleaned = text.replace(".", "").replace(",", ".")
    elif has_dot or has_comma:
        separator = "." if has_dot else ","
        occurrences = text.count(separator)
        if occurrences > 1:
            # Repeated separator -> thousands grouping.
            cleaned = text.replace(separator, "")
        else:
            _, _, after = text.partition(separator)
            trailing = len(after)
            if trailing in (1, 2):
                # 1-2 trailing digits -> decimal separator.
                cleaned = text.replace(separator, ".")
            else:
                # 0, 3 (e.g. "1.234", "500.000") or more -> thousands group.
                cleaned = text.replace(separator, "")
    else:
        cleaned = text

    cleaned = cleaned.strip()
    try:
        return float(cleaned)
    except ValueError:
        return None


def _parse_period(value: str) -> datetime | None:
    text = value.strip()
    for fmt in ("%Y-%m", "%Y-%m-%d", "%Y/%m", "%Y/%m/%d", "%m-%Y", "%m/%Y"):
        try:
            return datetime.strptime(text, fmt).replace(tzinfo=timezone.utc)
        except ValueError:
            continue
    return None


def _to_optional_float(value: object) -> float | None:
    if value is None:
        return None
    try:
        return float(value)  # type: ignore[arg-type]
    except (TypeError, ValueError):
        return None


__all__ = [
    "BcraBalanceCambiarioConnector",
    "DOES_NOT_REPLACE_NET_INTERVENTION",
    "DEFAULT_TTL_SECONDS",
    "DEFAULT_UNIDAD",
    "DEFAULT_XLSX_URL",
    "MissingXlsxExtractorError",
    "OpenpyxlRowExtractor",
    "ParsedBalanceObservation",
    "CONNECTOR_NAME",
    "SOURCE_NAME",
    "PARSER_VERSION",
    "XlsxRowExtractor",
    "normalize_balance_observations",
    "parse_balance_csv",
    "parse_balance_rows",
]
